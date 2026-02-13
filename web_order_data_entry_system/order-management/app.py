import csv
import io
import os
import re
import threading
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
PRODUCTS_FILE = os.path.join(DATA_DIR, 'products.csv')
CUSTOMERS_FILE = os.path.join(DATA_DIR, 'customers.csv')
ORDERS_FILE = os.path.join(DATA_DIR, 'orders.csv')

PRODUCTS_FIELDS = ['id', 'name', 'unit', 'price']
CUSTOMERS_FIELDS = ['id', 'name']
ORDERS_FIELDS = ['id', 'date', 'customer', 'product', 'unit', 'price', 'quantity', 'total']

lock = threading.Lock()


def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def read_csv(filepath, fields):
    if not os.path.exists(filepath):
        return []
    with open(filepath, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)


def write_csv(filepath, fields, rows):
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def next_id(rows):
    if not rows:
        return 1
    return max(int(r['id']) for r in rows) + 1


# --- Products ---

@app.route('/api/products', methods=['GET'])
def get_products():
    rows = read_csv(PRODUCTS_FILE, PRODUCTS_FIELDS)
    return jsonify(rows)


@app.route('/api/products', methods=['POST'])
def add_product():
    data = request.json
    if not data or not data.get('name'):
        return jsonify({'error': '货物名称不能为空'}), 400
    with lock:
        rows = read_csv(PRODUCTS_FILE, PRODUCTS_FIELDS)
        name = data['name'].strip()
        if any(r['name'] == name for r in rows):
            return jsonify({'error': '货物名称已存在'}), 400
        row = {
            'id': next_id(rows),
            'name': name,
            'unit': data.get('unit', '').strip(),
            'price': data.get('price', '0'),
        }
        rows.append(row)
        write_csv(PRODUCTS_FILE, PRODUCTS_FIELDS, rows)
    return jsonify(row), 201


@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    data = request.json
    if not data:
        return jsonify({'error': '无数据'}), 400
    with lock:
        rows = read_csv(PRODUCTS_FILE, PRODUCTS_FIELDS)
        for r in rows:
            if int(r['id']) == pid:
                if 'name' in data:
                    new_name = data['name'].strip()
                    if new_name != r['name'] and any(
                        x['name'] == new_name for x in rows if int(x['id']) != pid
                    ):
                        return jsonify({'error': '货物名称已存在'}), 400
                    r['name'] = new_name
                if 'unit' in data:
                    r['unit'] = data['unit'].strip()
                if 'price' in data:
                    r['price'] = data['price']
                write_csv(PRODUCTS_FILE, PRODUCTS_FIELDS, rows)
                return jsonify(r)
        return jsonify({'error': '货物不存在'}), 404


@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    with lock:
        rows = read_csv(PRODUCTS_FILE, PRODUCTS_FIELDS)
        new_rows = [r for r in rows if int(r['id']) != pid]
        if len(new_rows) == len(rows):
            return jsonify({'error': '货物不存在'}), 404
        write_csv(PRODUCTS_FILE, PRODUCTS_FIELDS, new_rows)
    return jsonify({'ok': True})


# --- Customers ---

@app.route('/api/customers', methods=['GET'])
def get_customers():
    rows = read_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS)
    return jsonify(rows)


@app.route('/api/customers', methods=['POST'])
def add_customer():
    data = request.json
    if not data or not data.get('name'):
        return jsonify({'error': '客户名称不能为空'}), 400
    with lock:
        rows = read_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS)
        name = data['name'].strip()
        if any(r['name'] == name for r in rows):
            return jsonify({'error': '客户名称已存在'}), 400
        row = {
            'id': next_id(rows),
            'name': name,
        }
        rows.append(row)
        write_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS, rows)
    return jsonify(row), 201


@app.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    data = request.json
    if not data:
        return jsonify({'error': '无数据'}), 400
    with lock:
        rows = read_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS)
        for r in rows:
            if int(r['id']) == cid:
                if 'name' in data:
                    new_name = data['name'].strip()
                    if new_name != r['name'] and any(
                        x['name'] == new_name for x in rows if int(x['id']) != cid
                    ):
                        return jsonify({'error': '客户名称已存在'}), 400
                    r['name'] = new_name
                write_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS, rows)
                return jsonify(r)
        return jsonify({'error': '客户不存在'}), 404


@app.route('/api/customers/<int:cid>', methods=['DELETE'])
def delete_customer(cid):
    with lock:
        rows = read_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS)
        new_rows = [r for r in rows if int(r['id']) != cid]
        if len(new_rows) == len(rows):
            return jsonify({'error': '客户不存在'}), 404
        write_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS, new_rows)
    return jsonify({'ok': True})


# --- Orders ---

@app.route('/api/orders', methods=['GET'])
def get_orders():
    rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
    return jsonify(rows)


@app.route('/api/orders', methods=['POST'])
def add_order():
    data = request.json
    if not data:
        return jsonify({'error': '无数据'}), 400
    required = ['date', 'customer', 'product', 'quantity']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field} 不能为空'}), 400
    price = float(data.get('price', 0))
    quantity = float(data.get('quantity', 0))
    with lock:
        rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
        row = {
            'id': next_id(rows),
            'date': data['date'].strip(),
            'customer': data['customer'].strip(),
            'product': data['product'].strip(),
            'unit': data.get('unit', '').strip(),
            'price': str(price),
            'quantity': str(quantity),
            'total': str(round(price * quantity, 2)),
        }
        rows.append(row)
        write_csv(ORDERS_FILE, ORDERS_FIELDS, rows)
    return jsonify(row), 201


@app.route('/api/orders/<int:oid>', methods=['PUT'])
def update_order(oid):
    data = request.json
    if not data:
        return jsonify({'error': '无数据'}), 400
    with lock:
        rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
        for r in rows:
            if int(r['id']) == oid:
                for field in ['date', 'customer', 'product', 'unit']:
                    if field in data:
                        r[field] = str(data[field]).strip()
                if 'price' in data:
                    r['price'] = str(data['price'])
                if 'quantity' in data:
                    r['quantity'] = str(data['quantity'])
                r['total'] = str(round(float(r['price']) * float(r['quantity']), 2))
                write_csv(ORDERS_FILE, ORDERS_FIELDS, rows)
                return jsonify(r)
        return jsonify({'error': '订单不存在'}), 404


@app.route('/api/orders/<int:oid>', methods=['DELETE'])
def delete_order(oid):
    with lock:
        rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
        new_rows = [r for r in rows if int(r['id']) != oid]
        if len(new_rows) == len(rows):
            return jsonify({'error': '订单不存在'}), 404
        write_csv(ORDERS_FILE, ORDERS_FIELDS, new_rows)
    return jsonify({'ok': True})


@app.route('/api/orders/search', methods=['GET'])
def search_orders():
    customer = request.args.get('customer', '')
    product = request.args.get('product', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
    results = []

    for r in rows:
        if customer:
            try:
                if not re.search(customer, r['customer']):
                    continue
            except re.error:
                if customer not in r['customer']:
                    continue
        if product:
            if product not in r['product']:
                continue
        if date_from:
            if r['date'] < date_from:
                continue
        if date_to:
            # 支持按年(YYYY)、按月(YYYY-MM)、按日(YYYY-MM-DD)查询
            # date_to 作为上界，需要包含当天/当月/当年
            if len(date_to) == 4:
                if r['date'][:4] > date_to:
                    continue
            elif len(date_to) == 7:
                if r['date'][:7] > date_to:
                    continue
            else:
                if r['date'] > date_to:
                    continue

        results.append(r)

    total = sum(float(r.get('total', 0)) for r in results)
    return jsonify({'orders': results, 'total': round(total, 2)})


# --- Export / Import ---

def filter_orders(rows):
    """Filter orders by query params, reusing search_orders logic."""
    customer = request.args.get('customer', '')
    product = request.args.get('product', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    results = []
    for r in rows:
        if customer:
            try:
                if not re.search(customer, r['customer']):
                    continue
            except re.error:
                if customer not in r['customer']:
                    continue
        if product:
            if product not in r['product']:
                continue
        if date_from:
            if r['date'] < date_from:
                continue
        if date_to:
            if len(date_to) == 4:
                if r['date'][:4] > date_to:
                    continue
            elif len(date_to) == 7:
                if r['date'][:7] > date_to:
                    continue
            else:
                if r['date'] > date_to:
                    continue
        results.append(r)
    return results


@app.route('/api/orders/export/csv')
def export_orders_csv():
    rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
    results = filter_orders(rows)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=ORDERS_FIELDS)
    writer.writeheader()
    writer.writerows(results)

    output = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(output, mimetype='text/csv', as_attachment=True,
                     download_name=f'orders_{ts}.csv')


@app.route('/api/orders/export/excel')
def export_orders_excel():
    rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
    results = filter_orders(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = '订单'
    headers = ['ID', '日期', '客户', '货物', '单位', '单价', '数量', '总价']
    ws.append(headers)
    for r in results:
        ws.append([r.get(f, '') for f in ORDERS_FIELDS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'orders_{ts}.xlsx')


@app.route('/api/data/export')
def export_all_data():
    wb = Workbook()

    # Products sheet
    ws_p = wb.active
    ws_p.title = '货物'
    ws_p.append(['ID', '名称', '单位', '单价'])
    for r in read_csv(PRODUCTS_FILE, PRODUCTS_FIELDS):
        ws_p.append([r.get(f, '') for f in PRODUCTS_FIELDS])

    # Customers sheet
    ws_c = wb.create_sheet('客户')
    ws_c.append(['ID', '名称'])
    for r in read_csv(CUSTOMERS_FILE, CUSTOMERS_FIELDS):
        ws_c.append([r.get(f, '') for f in CUSTOMERS_FIELDS])

    # Orders sheet
    ws_o = wb.create_sheet('订单')
    ws_o.append(['ID', '日期', '客户', '货物', '单位', '单价', '数量', '总价'])
    for r in read_csv(ORDERS_FILE, ORDERS_FIELDS):
        ws_o.append([r.get(f, '') for f in ORDERS_FIELDS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'all_data_{ts}.xlsx')


@app.route('/api/orders/import', methods=['POST'])
def import_orders():
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '请选择文件'}), 400

    filename = f.filename.lower()
    imported = []

    if filename.endswith('.csv'):
        text = f.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            imported.append(row)
    elif filename.endswith('.xlsx'):
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            imported.append(dict(zip(headers, row)))
        wb.close()
    else:
        return jsonify({'error': '仅支持 .csv 和 .xlsx 文件'}), 400

    # Map possible header names to internal field names
    header_map = {
        'ID': 'id', 'id': 'id',
        '日期': 'date', 'date': 'date',
        '客户': 'customer', 'customer': 'customer',
        '货物': 'product', 'product': 'product',
        '单位': 'unit', 'unit': 'unit',
        '单价': 'price', 'price': 'price',
        '数量': 'quantity', 'quantity': 'quantity',
        '总价': 'total', 'total': 'total',
    }

    count = 0
    with lock:
        rows = read_csv(ORDERS_FILE, ORDERS_FIELDS)
        for raw in imported:
            mapped = {}
            for k, v in raw.items():
                if k and k in header_map:
                    mapped[header_map[k]] = str(v).strip() if v is not None else ''
            # Validate required fields
            if not mapped.get('date') or not mapped.get('customer') or not mapped.get('product') or not mapped.get('quantity'):
                continue
            price = float(mapped.get('price', 0) or 0)
            quantity = float(mapped.get('quantity', 0) or 0)
            row = {
                'id': next_id(rows),
                'date': mapped['date'],
                'customer': mapped['customer'],
                'product': mapped['product'],
                'unit': mapped.get('unit', ''),
                'price': str(price),
                'quantity': str(quantity),
                'total': str(round(price * quantity, 2)),
            }
            rows.append(row)
            count += 1
        if count > 0:
            write_csv(ORDERS_FILE, ORDERS_FIELDS, rows)

    return jsonify({'ok': True, 'count': count})


# --- Main page ---

@app.route('/')
def index():
    return render_template('index.html')


if __name__ == '__main__':
    ensure_data_dir()
    app.run(debug=True, host='127.0.0.1', port=5001)
